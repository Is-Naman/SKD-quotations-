"""
Flask web application for quotation automation.
"""
import csv
import io
import json
import os
import re
import base64
import mimetypes
from datetime import datetime
from pathlib import Path
import zipfile
import subprocess
import tempfile
import shutil

try:
    import openpyxl
except Exception:
    openpyxl = None

from flask import Flask, render_template, request, jsonify, send_file
from werkzeug.utils import secure_filename

# Import quotation automation module
try:
    from quotation_automation import (
        load_catalog,
        parse_catalog_file,
        parse_enquiry,
        build_quote_rows,
        QUOTE_HEADERS,
        safe_float,
    )
except Exception as exc:
    print(f"Warning: quotation_automation import failed: {exc}")
    load_catalog = lambda path: []
    parse_catalog_file = lambda path: []
    parse_enquiry = lambda enquiry, catalog: []
    build_quote_rows = lambda parsed_items, catalog: []
    QUOTE_HEADERS = [
        "product_id",
        "product_name",
        "requirement",
        "requested_quantity",
        "unit",
        "unit_price",
        "discount",
        "total_price",
        "note",
    ]
    safe_float = lambda value, default=0.0: float(str(value).strip()) if str(value).strip() else default

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16MB max upload
app.config["UPLOAD_FOLDER"] = os.path.join(os.path.dirname(__file__), "uploads")
app.config["CATALOG_FOLDER"] = os.path.join(os.path.dirname(__file__), "catalogs")

# Create directories if they don't exist
os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
os.makedirs(app.config["CATALOG_FOLDER"], exist_ok=True)
os.makedirs(os.path.join(os.path.dirname(__file__), "quotations"), exist_ok=True)

ALLOWED_EXTENSIONS = {"csv", "json", "pdf", "jpg", "jpeg", "png", "tif", "tiff"}


def sync_gemini_files(api_key):
    """Sync all PDFs and JPEGs in the uploads directory to Gemini."""
    from google import genai
    client = genai.Client(api_key=api_key)
    
    # 1. Fetch currently uploaded files on Gemini servers
    try:
        remote_files = {f.display_name: f for f in client.files.list()}
    except Exception as e:
        print(f"Error listing Gemini files: {e}")
        remote_files = {}

    uploads_dir = os.path.join(os.path.dirname(__file__), "uploads")
    local_files = []
    if os.path.exists(uploads_dir):
        for fn in os.listdir(uploads_dir):
            if fn.lower().endswith((".pdf", ".jpg", ".jpeg", ".png")):
                local_files.append(fn)

    synced_refs = []
    for fn in local_files:
        filepath = os.path.join(uploads_dir, fn)
        if fn in remote_files:
            synced_refs.append(remote_files[fn])
        else:
            print(f"Uploading {fn} to Gemini...")
            try:
                uploaded_file = client.files.upload(file=filepath, config={'display_name': fn})
                synced_refs.append(uploaded_file)
            except Exception as e:
                print(f"Failed to upload {fn} to Gemini: {e}")
                
    return synced_refs

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in {"pdf", "jpg", "jpeg", "png", "txt", "xlsx"}

@app.route("/")
def index():
    """Render the main page."""
    pdf_files = []
    uploads_dir = os.path.join(os.path.dirname(__file__), "uploads")
    if os.path.exists(uploads_dir):
        pdf_files = [f for f in os.listdir(uploads_dir) if f.lower().endswith((".pdf", ".jpeg", ".jpg", ".png"))]
    return render_template("index.html", catalogs=pdf_files)



def parse_json_robust(text):
    """Parse JSON string robustly, cleaning up markdown code block wrappers if present."""
    import json
    import re
    cleaned = text.strip()
    if cleaned.startswith("```"):
        match = re.search(r"```(?:json)?\s*(.*?)\s*```", cleaned, re.DOTALL | re.IGNORECASE)
        if match:
            cleaned = match.group(1).strip()
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        match = re.search(r"(\[.*\])", cleaned, re.DOTALL)
        if match:
            try:
                return json.loads(match.group(1))
            except json.JSONDecodeError:
                pass
        raise

@app.route("/api/generate-quotation", methods=["POST"])
def generate_quotation():
    """Generate a quotation from an enquiry using Gemini Multimodal Agent."""
    data = request.get_json()
 
    if not data or "enquiry" not in data:
        return jsonify({"error": "Missing enquiry text"}), 400
 
    enquiry_text = data.get("enquiry", "").strip()
    gemini_key = data.get("gemini_api_key", "").strip()
 
    if not enquiry_text:
        return jsonify({"error": "Enquiry text cannot be empty"}), 400
        
    if not gemini_key:
        return jsonify({"error": "Gemini API key is required for the AI agent"}), 400
 
    try:
        # 1. Sync files to Gemini
        file_refs = sync_gemini_files(gemini_key)
        
        # 2. Call Gemini 1.5 Pro to process the whole thing
        from google import genai
        client = genai.Client(api_key=gemini_key)
        
        prompt = f"""
You are an expert electrical materials estimation agent working for a distributor/contractor.
I have uploaded multiple manufacturer price lists, catalogs, and formula sheets as PDFs and JPEGs.

RULES — READ CAREFULLY:

RULE 1 — VAGUE / TYPE-ONLY ENQUIRIES (most important):
If the customer names a product TYPE without specifying an exact size, rating, or model
(e.g. "copper sector lugs long barrel", "MCB", "switches", "lugs", "glands"),
you MUST scan ALL uploaded catalogs thoroughly and list EVERY matching variant as a
separate row. Do NOT return an empty list or a single generic row.
Examples of what "every variant" means:
- "copper sector lugs" → list each size: 16sqmm, 25sqmm, 35sqmm, 50sqmm, 70sqmm, 95sqmm,
  120sqmm, 150sqmm, 185sqmm, 240sqmm, 300sqmm etc., across all brands found in catalogs.
- "MCB" → list each ampere rating: 6A, 10A, 16A, 20A, 25A, 32A, 40A, 63A etc.
This is a product EXPLORATION request — give every option. The user will delete what they don't need.

RULE 2 — SPECIFIC ENQUIRIES:
If the customer specifies quantities and exact specifications, match each line item exactly.
One catalog match = one row.

RULE 3 — QUANTITY:
If no quantity is stated, default requested_quantity to 1 for each variant row.

RULE 4 — PRICING:
Read prices ONLY from the uploaded catalog PDFs/images. Never guess or estimate.
If a price is given as a formula (e.g. base price × factor), compute it.

RULE 5 — NOT FOUND:
If after thoroughly searching all catalogs a product is genuinely not found,
set unit_price to "Regret" and total_price to 0.
Never return an empty array — at minimum return a "Regret" row.

RULE 6 — SOURCE TRACING (critical):
For every row you output, you MUST record exactly which catalog file and page you read
the price from. This is non-negotiable — the user needs to be able to verify every price.
  "source_file" — the exact filename of the PDF or image you read the price from
                  (e.g. "DOWELL'S 25-05-2026 NEW LP.pdf" or "Schneider LP April 26.jpeg").
                  If you could not find the product, write "Not found".
  "source_page" — the page number within that file where the price appears (integer).
                  For single-page images (JPEG), always write 1.
                  If not found, write 0.

RULE 7 — OUTPUT FORMAT (strict JSON array):
Each element must have EXACTLY these keys:
  "product_id"         — catalog code/part number (blank string if none)
  "product_name"       — exact name from catalog including size/rating/type
  "requirement"        — copy of the original enquiry description for this item
  "requested_quantity" — integer (1 if not stated)
  "unit"               — "Nos", "Mtr", "Set", "Kg", etc.
  "unit_price"         — numeric list price from catalog, OR "Regret" if not found
  "discount"           — "0%" unless you know the applicable discount
  "total_price"        — requested_quantity × unit_price × (1 - discount_pct/100), or 0
  "note"               — brand name + key spec (e.g. "Dowells — Long Barrel, Tin Plated, 2-Hole")
  "source_file"        — filename of the catalog PDF/JPEG this price was taken from
  "source_page"        — page number within that file (integer, 1 for single-page images)

Customer Enquiry:
{enquiry_text}

Return ONLY a valid JSON array. No markdown fences, no explanation text, just the array.
"""

        contents = [*file_refs, prompt]
        
        import time
        start = time.time()
        print("Sending multimodal request to Gemini (this may take up to a minute)...")
        # In new SDK, we use models.generate_content
        # Note: the new SDK accepts parts directly in the contents list, but if they are objects, we pass them as a list.
        response = client.models.generate_content(
            model='gemini-2.5-flash',
            contents=contents,
        )
        print(f"Gemini responded in {time.time() - start:.1f}s")
        
        # 3. Parse JSON response
        quotation_data = parse_json_robust(response.text)
        
        # Fallback if it returned a dict instead of list
        if isinstance(quotation_data, dict) and "quotation" in quotation_data:
            quotation_data = quotation_data["quotation"]
        elif isinstance(quotation_data, dict) and "items" in quotation_data:
            quotation_data = quotation_data["items"]
            
        if not isinstance(quotation_data, list):
            quotation_data = [quotation_data]
            
        # 4. Save the quotation for later review
        try:
            from datetime import datetime
            quote_id = f"quote_{int(time.time())}"
            
            # Determine source details
            email_subject = data.get("email_subject")
            email_from = data.get("email_from")
            customer_name = data.get("customer_name") or ""
            
            source_str = "Manual Input"
            if email_subject:
                source_str = f"Email: {email_subject}"
                if email_from:
                    source_str += f" (From: {email_from})"
                    
            quote_save_data = {
                "id": quote_id,
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "customer_name": customer_name,
                "source": source_str,
                "enquiry": enquiry_text,
                "quotation": quotation_data,
                "total_amount": sum(float(r.get("total_price") or 0.0) if isinstance(r.get("total_price"), (int, float)) else 0.0 for r in quotation_data)
            }
            
            quote_dir = os.path.join(os.path.dirname(__file__), "quotations")
            os.makedirs(quote_dir, exist_ok=True)
            quote_path = os.path.join(quote_dir, f"{quote_id}.json")
            with open(quote_path, "w", encoding="utf-8") as f:
                json.dump(quote_save_data, f, indent=2)
        except Exception as save_err:
            print(f"Failed to auto-save quotation: {save_err}")

        return jsonify({
            "message": "Quotation generated successfully",
            "quotation": quotation_data,
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Failed to generate quotation: {str(e)}"}), 400


@app.route("/api/export-quotation", methods=["POST"])
def export_quotation():
    """Export quotation to CSV."""
    data = request.get_json()

    if not data or "quotation" not in data:
        return jsonify({"error": "No quotation provided"}), 400

    try:
        quotation = data.get("quotation", [])
        customer_name = data.get("customer_name", "Customer").replace(" ", "_")

        # Create CSV
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=QUOTE_HEADERS)
        writer.writeheader()

        for item in quotation:
            writer.writerow({
                "product_id": item.get("product_id", ""),
                "product_name": item.get("product_name", ""),
                "requirement": item.get("requirement", ""),
                "requested_quantity": item.get("requested_quantity", 1),
                "unit": item.get("unit", ""),
                "unit_price": item.get("unit_price", 0),
                "discount": item.get("discount", "0%"),
                "total_price": item.get("total_price", 0),
                "note": item.get("note", ""),
            })

        # Convert to bytes
        output.seek(0)
        output_bytes = output.getvalue().encode("utf-8")

        # Create response
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"quotation_{customer_name}_{timestamp}.csv"

        return send_file(
            io.BytesIO(output_bytes),
            mimetype="text/csv",
            as_attachment=True,
            download_name=filename,
        )

    except Exception as e:
        return jsonify({"error": f"Failed to export quotation: {str(e)}"}), 400


@app.route("/api/update-row", methods=["POST"])
def update_row():
    """Update a single quotation row (discount, quantity, etc.)."""
    data = request.get_json()

    if not data or "row_index" not in data:
        return jsonify({"error": "Missing row index"}), 400

    try:
        row_index = int(data.get("row_index"))
        updates = data.get("updates", {})

        # Validate and parse updates
        if "discount" in updates:
            discount_str = str(updates["discount"]).strip().rstrip("%")
            try:
                discount_pct = float(discount_str)
            except ValueError:
                return jsonify({"error": "Invalid discount value"}), 400
            updates["discount"] = f"{discount_pct}%"

        if "requested_quantity" in updates:
            updates["requested_quantity"] = int(updates["requested_quantity"])

        return jsonify({
            "message": "Row updated",
            "row_index": row_index,
            "updates": updates,
        }), 200

    except Exception as e:
        return jsonify({"error": f"Failed to update row: {str(e)}"}), 400


def run_local_ocr(image_path):
    """Run local WinRT OCR via PowerShell wrapper script."""
    script_path = os.path.join(os.path.dirname(__file__), "ocr_helper.ps1")
    cmd = ["powershell", "-ExecutionPolicy", "Bypass", "-File", script_path, image_path]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, check=True)
        return result.stdout.strip()
    except Exception as e:
        print(f"OCR failed for {image_path}: {e}")
        return ""


def run_gemini_vision_ocr(image_path, api_key):
    """Use Gemini Vision API to accurately extract text/tables from an image."""
    try:
        from google import genai
        from PIL import Image
        
        client = genai.Client(api_key=api_key)
        
        img = Image.open(image_path)
        prompt = "Extract all text and tables from this image perfectly line by line. Maintain row alignment for tables. Just return the extracted text, no markdown wrappers, no explanations."
        
        response = client.models.generate_content(
            model='gemini-2.5-flash',
            contents=[img, prompt]
        )
        return response.text.strip()
    except Exception as e:
        print(f"Gemini Vision API SDK failed: {e}")
        return run_local_ocr(image_path)


def parse_excel_enquiry(file_path):
    """Parse text from cells and extract/OCR embedded images from an Excel file."""
    if openpyxl is None:
        return "Excel parsing is unavailable because openpyxl is not installed."

    text_content = []
    
    # 1. Read sheet cell text using openpyxl
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_lines = []
            for row in ws.iter_rows(values_only=True):
                # Format row cells
                row_vals = [str(val).strip() for val in row if val is not None]
                if row_vals:
                    sheet_lines.append(" \t ".join(row_vals))
            if sheet_lines:
                text_content.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(sheet_lines))
    except Exception as e:
        print(f"Failed to read cells from Excel: {e}")
        text_content.append(f"Error reading cells: {e}")

    # 2. Extract and OCR embedded images from xlsx zip structure
    try:
        temp_dir = tempfile.mkdtemp()
        extracted_images = []
        with zipfile.ZipFile(file_path, 'r') as zf:
            for name in zf.namelist():
                if name.startswith("xl/media/image"):
                    # Extract the image
                    dest_path = os.path.join(temp_dir, os.path.basename(name))
                    with zf.open(name) as source, open(dest_path, 'wb') as target:
                        shutil.copyfileobj(source, target)
                    extracted_images.append(dest_path)
        
        # Run OCR on all extracted images
        ocr_texts = []
        for img_idx, img_path in enumerate(extracted_images, start=1):
            text = run_local_ocr(img_path)
            if text:
                ocr_texts.append(f"[Image {img_idx} OCR]:\n{text}")
        
        if ocr_texts:
            text_content.append("\n--- Extracted Image Text ---\n" + "\n\n".join(ocr_texts))
            
        # Clean up temp folder
        shutil.rmtree(temp_dir, ignore_errors=True)
    except Exception as e:
        print(f"Failed to extract images from Excel: {e}")
        
    return "\n\n".join(text_content)


@app.route("/api/fetch-email", methods=["POST"])
def fetch_email():
    """Fetch the latest RFQ/enquiry email body and attachments via IMAP."""
    import imaplib
    import email
    from email.header import decode_header
    
    data = request.get_json()
    if not data:
        return jsonify({"error": "Missing config data"}), 400
        
    imap_server = data.get("imap_server", "imap.gmail.com").strip()
    email_address = data.get("email", "").strip()
    password = data.get("password", "").strip()
    subject_filter = data.get("subject_filter", "RFQ, Enquiry").strip()
    
    if not email_address or not password:
        return jsonify({"error": "Email address and password/app-password are required"}), 400
        
    try:
        # Connect to IMAP
        mail = imaplib.IMAP4_SSL(imap_server, 993)
        mail.login(email_address, password)
        mail.select("inbox")
        
        # Search for messages
        status, messages = mail.search(None, "ALL")
        if status != "OK":
            return jsonify({"error": "Failed to search inbox"}), 400
            
        mail_ids = messages[0].split()
        if not mail_ids:
            return jsonify({"error": "No emails found in inbox"}), 404
            
        # Walk backwards from the newest email
        latest_text = ""
        email_subject = ""
        email_from = ""
        found_match = False
        
        # Filter terms
        filter_terms = [t.strip().lower() for t in subject_filter.split(",") if t.strip()]
        
        for mail_id in reversed(mail_ids):
            # Fetch email headers first to check subject
            res_status, data_parts = mail.fetch(mail_id, "(RFC822)")
            if res_status != "OK":
                continue
                
            raw_email = data_parts[0][1]
            msg = email.message_from_bytes(raw_email)
            
            # Decode subject
            subject, encoding = decode_header(msg["Subject"] or "")[0]
            if isinstance(subject, bytes):
                subject = subject.decode(encoding or "utf-8", errors="ignore")
                
            subject_lower = subject.lower()
            
            # Check if matches filter
            if filter_terms:
                if not any(term in subject_lower for term in filter_terms):
                    continue
            
            # Found matching email!
            found_match = True
            email_subject = subject
            email_from = msg["From"]
            
            # Extract body
            body_parts = []
            attachment_text = []
            
            if msg.is_multipart():
                for part in msg.walk():
                    content_type = part.get_content_type()
                    content_disposition = str(part.get("Content-Disposition"))
                    
                    # Extract text body
                    if content_type == "text/plain" and "attachment" not in content_disposition:
                        payload = part.get_payload(decode=True)
                        if payload:
                            body_parts.append(payload.decode("utf-8", errors="ignore"))
                            
                    # Extract attachments (Excel, PDF, Text)
                    elif "attachment" in content_disposition:
                        filename = part.get_filename()
                        if filename:
                            filename, encoding = decode_header(filename)[0]
                            if isinstance(filename, bytes):
                                filename = filename.decode(encoding or "utf-8", errors="ignore")
                            
                            ext = Path(filename).suffix.lower()
                            if ext in {".xlsx", ".xls", ".txt"}:
                                # Save attachment temporarily to parse it
                                temp_dir = os.path.join(app.config["UPLOAD_FOLDER"], "temp_email")
                                os.makedirs(temp_dir, exist_ok=True)
                                temp_path = os.path.join(temp_dir, filename)
                                with open(temp_path, "wb") as f:
                                    f.write(part.get_payload(decode=True))
                                    
                                if ext in {".xlsx", ".xls"}:
                                    text_extracted = parse_excel_enquiry(temp_path)
                                    attachment_text.append(f"--- Attachment: {filename} ---\n{text_extracted}")
                                elif ext == ".txt":
                                    with open(temp_path, "r", encoding="utf-8", errors="ignore") as tf:
                                        attachment_text.append(f"--- Attachment: {filename} ---\n{tf.read()}")
                                        
                                os.remove(temp_path)
            else:
                payload = msg.get_payload(decode=True)
                if payload:
                    body_parts.append(payload.decode("utf-8", errors="ignore"))
                    
            latest_text = "\n\n".join(body_parts).strip()
            if attachment_text:
                latest_text += "\n\n" + "\n\n".join(attachment_text)
                
            break
            
        mail.close()
        mail.logout()
        
        if not found_match:
            return jsonify({"error": f"No emails matching filter '{subject_filter}' found."}), 404
            
        return jsonify({
            "message": "Email fetched successfully",
            "subject": email_subject,
            "from": email_from,
            "body": latest_text
        })
        
    except imaplib.IMAP4.error as e:
        return jsonify({"error": f"IMAP Login failed: {str(e)}. Please check your credentials or verify if your mail server requires an App Password."}), 401
    except Exception as e:
        return jsonify({"error": f"Failed to fetch email: {str(e)}"}), 500


@app.route("/api/saved-quotations", methods=["GET"])
def list_saved_quotations():
    """List all saved quotations for review."""
    import glob
    quote_dir = os.path.join(os.path.dirname(__file__), "quotations")
    os.makedirs(quote_dir, exist_ok=True)
    
    saved_quotes = []
    for filepath in glob.glob(os.path.join(quote_dir, "*.json")):
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                qdata = json.load(f)
                saved_quotes.append({
                    "id": qdata.get("id"),
                    "timestamp": qdata.get("timestamp"),
                    "customer_name": qdata.get("customer_name"),
                    "source": qdata.get("source"),
                    "total_amount": qdata.get("total_amount", 0.0)
                })
        except Exception as e:
            print(f"Error loading saved quote file {filepath}: {e}")
            
    # Sort by timestamp descending
    saved_quotes.sort(key=lambda x: x.get("timestamp", ""), reverse=True)
    return jsonify({"saved_quotations": saved_quotes})


@app.route("/api/saved-quotations/<quote_id>", methods=["GET"])
def get_saved_quotation(quote_id):
    """Retrieve full details of a saved quotation."""
    quote_dir = os.path.join(os.path.dirname(__file__), "quotations")
    quote_path = os.path.join(quote_dir, f"{secure_filename(quote_id)}.json")
    if not os.path.exists(quote_path):
        return jsonify({"error": "Saved quotation not found"}), 404
        
    try:
        with open(quote_path, "r", encoding="utf-8") as f:
            qdata = json.load(f)
        return jsonify(qdata)
    except Exception as e:
        return jsonify({"error": f"Failed to read saved quotation: {str(e)}"}), 500


@app.route("/api/saved-quotations/<quote_id>", methods=["DELETE"])
def delete_saved_quotation(quote_id):
    """Delete a saved quotation."""
    quote_dir = os.path.join(os.path.dirname(__file__), "quotations")
    quote_path = os.path.join(quote_dir, f"{secure_filename(quote_id)}.json")
    if not os.path.exists(quote_path):
        return jsonify({"error": "Saved quotation not found"}), 404
        
    try:
        os.remove(quote_path)
        return jsonify({"message": "Saved quotation deleted successfully"}), 200
    except Exception as e:
        return jsonify({"error": f"Failed to delete saved quotation: {str(e)}"}), 500


@app.route("/api/parse-enquiry-file", methods=["POST"])
def parse_enquiry_file():
    """Parse enquiry text from uploaded file (Excel, Image, or Text)."""
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No file selected"}), 400

    try:
        gemini_api_key = request.form.get("gemini_api_key", "").strip()
        filename = secure_filename(file.filename)
        enquiry_dir = os.path.join(app.config["UPLOAD_FOLDER"], "enquiries")
        os.makedirs(enquiry_dir, exist_ok=True)
        filepath = os.path.join(enquiry_dir, filename)
        file.save(filepath)

        ext = Path(filename).suffix.lower()
        extracted_text = ""

        if ext in {".xlsx", ".xls"}:
            extracted_text = parse_excel_enquiry(filepath)
        elif ext in {".png", ".jpg", ".jpeg", ".tif", ".tiff"}:
            if gemini_api_key:
                extracted_text = run_gemini_vision_ocr(filepath, gemini_api_key)
            else:
                extracted_text = run_local_ocr(filepath)
        elif ext in {".txt"}:
            with open(filepath, "r", encoding="utf-8", errors="ignore") as fh:
                extracted_text = fh.read()
        else:
            return jsonify({"error": "Unsupported file type. Please upload Excel (.xlsx), Image (.png, .jpg), or Text (.txt)."}), 400

        # Clean up uploaded file
        try:
            os.remove(filepath)
        except Exception:
            pass

        return jsonify({
            "message": "Enquiry file parsed successfully",
            "text": extracted_text,
        }), 200

    except Exception as e:
        return jsonify({"error": f"Failed to parse enquiry file: {str(e)}"}), 400

@app.route("/api/saved-quotes", methods=["GET"])
def list_saved_quotes_alias():
    """Alias: list quotation files for the chat UI sidebar."""
    import glob
    quote_dir = os.path.join(os.path.dirname(__file__), "quotations")
    os.makedirs(quote_dir, exist_ok=True)
    files = []
    for fp in sorted(glob.glob(os.path.join(quote_dir, "*.json")), reverse=True):
        files.append({"name": os.path.basename(fp)})
    return jsonify({"files": files})


@app.route("/api/download-quote/<filename>", methods=["GET"])
def download_quote_file(filename):
    """Download a saved quotation JSON file by filename."""
    quote_dir = os.path.join(os.path.dirname(__file__), "quotations")
    safe_name = secure_filename(filename)
    quote_path = os.path.join(quote_dir, safe_name)
    if not os.path.exists(quote_path):
        return jsonify({"error": "File not found"}), 404
    return send_file(quote_path, as_attachment=True, download_name=safe_name)


if __name__ == "__main__":
    app.run(debug=True, use_reloader=False, host="0.0.0.0", port=5000)

# Vercel-compatible WSGI entry point
application = app
